# analytics/views.py

from django.shortcuts import render
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status, permissions
from integrations.services.prediction_service import PredictionService
from django.db.models import Count, Avg, Max, Min, Q, Sum, F, ExpressionWrapper, fields
from django.db.models.functions import TruncHour, TruncDate, ExtractHour, ExtractWeekDay, Extract
from django.utils import timezone
from datetime import datetime, timedelta
from django.core.cache import cache
import json
import csv
import io
from django.http import HttpResponse
from django.template.loader import get_template
import logging

# Excel support
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF support
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from nfc_hospital_system.utils import APIResponse
from authentication.models import User # 이 User는 커스텀 User 모델입니다.
from nfc.models import NFCTag, TagLog
from p_queue.models import Queue, QueueStatusLog
from appointments.models import Appointment, Exam
from admin_dashboard.models import AdminLog

logger = logging.getLogger(__name__)

# 권한 확인 헬퍼 함수
def _check_analytics_permission(request):
    """분석/통계 권한 확인 - Dept-Admin 이상"""
    # 데모 모드에서는 권한 체크 완화
    if cache.get("demo_mode_active"):
        return True, None

    # request.user는 이미 인증된 User 객체 (또는 AnonymousUser)입니다.
    # User 모델을 다시 쿼리할 필요 없이 request.user 객체의 role을 바로 확인합니다.
    if not request.user.is_authenticated:
        return False, "로그인이 필요합니다."

    # request.user는 authentication.models.User 인스턴스라고 가정
    # admin_user = User.objects.get(user=request.user) # 이 줄이 문제였습니다. 제거하거나 주석 처리합니다.

    # 직접 request.user.role을 확인합니다.
    # 테스트를 위해 임시로 모든 인증된 사용자 허용
    # if request.user.role not in ['super', 'dept', 'staff', 'patient']:
    #     return False, "부서 관리자 이상의 권한이 필요합니다."

    return True, None  # 임시로 모든 인증된 사용자 허용

# 통계 데이터 API

@api_view(['GET'])
@permission_classes([])  # 테스트를 위해 임시로 권한 제거
def patient_flow_analysis(request):
    """
    환자 동선 분석 API - GET /analytics/patient-flow
    
    NFC 태그 스캔 로그를 기반으로 환자 이동 패턴 분석
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 쿼리 파라미터
        start_date = request.GET.get('startDate', (timezone.now() - timedelta(days=7)).isoformat())
        end_date = request.GET.get('endDate', timezone.now().isoformat())
        department = request.GET.get('department')
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # 태그 로그 기반 환자 동선 분석
        tag_logs = TagLog.objects.filter(
            timestamp__range=[start_date, end_date]
        ).select_related('tag', 'user')
        
        # 가장 많이 방문한 위치
        popular_locations = tag_logs.values(
            'tag__building', 'tag__floor', 'tag__room'
        ).annotate(
            visit_count=Count('log_id'),
            unique_visitors=Count('user', distinct=True)
        ).order_by('-visit_count')[:10]
        
        # 위치 정보를 문자열로 조합
        popular_locations_formatted = []
        for loc in popular_locations:
            popular_locations_formatted.append({
                'location': f"{loc['tag__building']} {loc['tag__floor']}층 {loc['tag__room']}",
                'visit_count': loc['visit_count'],
                'unique_visitors': loc['unique_visitors']
            })
        
        # 위치별 평균 체류 시간 (다음 스캔까지의 시간)
        location_durations = {}
        users = tag_logs.values_list('user', flat=True).distinct()
        
        for user_id in users:
            user_logs = tag_logs.filter(user_id=user_id).order_by('timestamp')
            for i in range(len(user_logs) - 1):
                current_log = user_logs[i]
                next_log = user_logs[i + 1]
                location = current_log.tag.get_location_display()
                duration = (next_log.timestamp - current_log.timestamp).total_seconds() / 60
                
                if location not in location_durations:
                    location_durations[location] = []
                location_durations[location].append(duration)
        
        # 평균 계산
        avg_durations = {}
        for location, durations in location_durations.items():
            if durations:
                avg_durations[location] = {
                    'avgDuration': round(sum(durations) / len(durations), 2),
                    'maxDuration': round(max(durations), 2),
                    'minDuration': round(min(durations), 2)
                }
        
        # 이동 경로 패턴 분석
        flow_patterns = []
        for user_id in users[:100]:  # 최대 100명까지만 분석
            user_logs = tag_logs.filter(user_id=user_id).order_by('timestamp')
            if user_logs.count() >= 2:
                path = []
                for log in user_logs:
                    path.append({
                        'location': log.tag.get_location_display(),
                        'timestamp': log.timestamp.isoformat(),
                        'tagCode': log.tag.code
                    })
                flow_patterns.append({
                    'userId': str(user_id),
                    'path': path,
                    'totalStops': len(path)
                })
        
        # 시간대별 혼잡도
        hourly_traffic = tag_logs.annotate(
            hour=ExtractHour('timestamp')
        ).values('hour').annotate(
            scan_count=Count('log_id'),
            unique_users=Count('user', distinct=True)
        ).order_by('hour')
        
        # 병목 구간 식별 (같은 위치에서 짧은 시간 내 여러 스캔)
        bottlenecks = []
        location_scans = tag_logs.values('tag__building', 'tag__floor', 'tag__room').annotate(
            total_scans=Count('log_id'),
            unique_users=Count('user', distinct=True),
            scans_per_user=F('total_scans') / F('unique_users')
        ).filter(scans_per_user__gt=3).order_by('-scans_per_user')
        
        for loc in location_scans:
            bottlenecks.append({
                'location': f"{loc['tag__building']} {loc['tag__floor']}층 {loc['tag__room']}",
                'totalScans': loc['total_scans'],
                'uniqueUsers': loc['unique_users'],
                'avgScansPerUser': round(loc['scans_per_user'], 2),
                'congestionLevel': 'high' if loc['scans_per_user'] > 5 else 'medium'
            })
        
        return APIResponse.success(
            data={
                'summary': {
                    'totalScans': tag_logs.count(),
                    'uniquePatients': tag_logs.values('user').distinct().count(),
                    'dateRange': {
                        'start': start_date.isoformat(),
                        'end': end_date.isoformat()
                    }
                },
                'popularLocations': popular_locations_formatted,
                'locationDurations': avg_durations,
                'flowPatterns': flow_patterns[:20],  # 최대 20개 패턴만 반환
                'hourlyTraffic': list(hourly_traffic),
                'bottlenecks': bottlenecks
            },
            message="환자 동선 분석을 완료했습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"Patient flow analysis error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="환자 동선 분석 중 오류가 발생했습니다.",
            code="FLOW_ANALYSIS_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([])  # 테스트를 위해 임시로 권한 제거
def waiting_time_statistics(request):
    """
    대기시간 통계 API - GET /analytics/waiting-time
    
    대기시간 통계 및 트렌드 분석
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 쿼리 파라미터
        start_date = request.GET.get('startDate', (timezone.now() - timedelta(days=7)).isoformat())
        end_date = request.GET.get('endDate', timezone.now().isoformat())
        department = request.GET.get('department')
        granularity = request.GET.get('granularity', 'daily')  # hourly, daily, weekly, monthly
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # 대기열 데이터 필터링
        # `joined_at` 필드 대신 `created_at` 필드를 사용
        queues = Queue.objects.filter(
            created_at__range=[start_date, end_date]
        )
        
        if department:
            queues = queues.filter(exam__department=department)
        
        # 기본 통계
        basic_stats = queues.aggregate(
            avg_wait_time=Avg('estimated_wait_time'),
            max_wait_time=Max('estimated_wait_time'),
            min_wait_time=Min('estimated_wait_time'),
            total_patients=Count('queue_id')
        )
        
        # 실제 대기시간 vs 예상 대기시간
        completed_queues = queues.filter(
            state='completed',
            called_at__isnull=False
        )
        
        wait_time_comparison = []
        for q in completed_queues:
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            if q.called_at and q.created_at:
                actual_wait = (q.called_at - q.created_at).total_seconds() / 60
                estimated_wait = q.estimated_wait_time or 0
                wait_time_comparison.append({
                    'actual': round(actual_wait, 2),
                    'estimated': round(estimated_wait, 2),
                    'difference': round(actual_wait - estimated_wait, 2),
                    'accuracy': round((1 - abs(actual_wait - estimated_wait) / estimated_wait) * 100, 2) if estimated_wait > 0 else 0
                })
        
        # 시간대별 트렌드
        if granularity == 'hourly':
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            time_trends = queues.annotate(
                period=ExtractHour('created_at')
            ).values('period').annotate(
                avg_wait=Avg('estimated_wait_time'),
                patient_count=Count('queue_id')
            ).order_by('period')
        elif granularity == 'weekly':
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            time_trends = queues.annotate(
                period=ExtractWeekDay('created_at')
            ).values('period').annotate(
                avg_wait=Avg('estimated_wait_time'),
                patient_count=Count('queue_id')
            ).order_by('period')
        elif granularity == 'monthly':
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            time_trends = queues.annotate(
                period=TruncDate('created_at')
            ).values('period').annotate(
                avg_wait=Avg('estimated_wait_time'),
                patient_count=Count('queue_id')
            ).order_by('period')
        else:  # daily
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            time_trends = queues.annotate(
                period=TruncDate('created_at')
            ).values('period').annotate(
                avg_wait=Avg('estimated_wait_time'),
                patient_count=Count('queue_id')
            ).order_by('period')
        
        # 우선순위별 대기시간
        priority_stats = queues.values('priority').annotate(
            avg_wait=Avg('estimated_wait_time'),
            max_wait=Max('estimated_wait_time'),
            min_wait=Min('estimated_wait_time'),
            count=Count('queue_id')
        )
        
        # 부서별 대기시간
        dept_stats = queues.values('exam__department').annotate(
            avg_wait=Avg('estimated_wait_time'),
            max_wait=Max('estimated_wait_time'),
            patient_count=Count('queue_id')
        ).order_by('-avg_wait')
        
        # 대기시간 분포
        wait_distribution = {
            'under_10min': queues.filter(estimated_wait_time__lt=10).count(),
            '10_to_30min': queues.filter(estimated_wait_time__gte=10, estimated_wait_time__lt=30).count(),
            '30_to_60min': queues.filter(estimated_wait_time__gte=30, estimated_wait_time__lt=60).count(),
            'over_60min': queues.filter(estimated_wait_time__gte=60).count()
        }
        
        return APIResponse.success(
            data={
                'summary': basic_stats,
                'waitTimeComparison': {
                    'samples': wait_time_comparison[:100],  # 최대 100개 샘플
                    'avgAccuracy': round(sum(w['accuracy'] for w in wait_time_comparison) / len(wait_time_comparison), 2) if wait_time_comparison else 0
                },
                'trends': list(time_trends),
                'priorityBreakdown': list(priority_stats),
                'departmentStats': list(dept_stats),
                'distribution': wait_distribution,
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat(),
                    'granularity': granularity
                }
            },
            message="대기시간 통계를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"Waiting time statistics error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="대기시간 통계 조회 중 오류가 발생했습니다.",
            code="WAIT_TIME_STATS_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def congestion_heatmap(request):
    """
    시간대별 혼잡도 히트맵 API - GET /analytics/congestion-heatmap
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 쿼리 파라미터
        start_date = request.GET.get('startDate', (timezone.now() - timedelta(days=7)).isoformat())
        end_date = request.GET.get('endDate', timezone.now().isoformat())
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # 시간대별, 위치별 혼잡도 데이터
        heatmap_data = []
        
        # 각 위치별로 시간대별 혼잡도 계산
        locations = NFCTag.objects.values('building', 'floor', 'room').distinct()
        
        for loc in locations:
            location = f"{loc['building']} {loc['floor']}층 {loc['room']}"
            location_data = {
                'location': location,
                'hourlyData': []
            }
            
            for hour in range(24):
                # 해당 시간대의 스캔 수 계산
                scans = TagLog.objects.filter(
                    tag__building=loc['building'],
                    tag__floor=loc['floor'],
                    tag__room=loc['room'],
                    timestamp__range=[start_date, end_date],
                    timestamp__hour=hour
                ).count()
                
                # 혼잡도 레벨 계산 (0-4)
                if scans == 0:
                    level = 0
                elif scans < 10:
                    level = 1
                elif scans < 30:
                    level = 2
                elif scans < 50:
                    level = 3
                else:
                    level = 4
                
                location_data['hourlyData'].append({
                    'hour': hour,
                    'scanCount': scans,
                    'congestionLevel': level
                })
            
            heatmap_data.append(location_data)
        
        # 요일별 패턴
        weekday_patterns = []
        for weekday in range(7):  # 0=Monday, 6=Sunday
            day_scans = TagLog.objects.filter(
                timestamp__range=[start_date, end_date],
                timestamp__week_day=weekday + 1  # Django uses 1=Sunday, 7=Saturday
            ).annotate(
                hour=ExtractHour('timestamp')
            ).values('hour').annotate(
                scan_count=Count('log_id')
            ).order_by('hour')
            
            weekday_patterns.append({
                'weekday': weekday,
                'hourlyScans': list(day_scans)
            })
        
        # 가장 혼잡한 시간대 TOP 10
        peak_times = TagLog.objects.filter(
            timestamp__range=[start_date, end_date]
        ).annotate(
            hour=ExtractHour('timestamp')
        ).values('hour', 'tag__building', 'tag__floor', 'tag__room').annotate(
            scan_count=Count('log_id')
        ).order_by('-scan_count')[:10]
        
        # 위치 정보를 문자열로 조합
        peak_times_formatted = []
        for pt in peak_times:
            peak_times_formatted.append({
                'hour': pt['hour'],
                'location': f"{pt['tag__building']} {pt['tag__floor']}층 {pt['tag__room']}",
                'scan_count': pt['scan_count']
            })
        
        return APIResponse.success(
            data={
                'heatmapData': heatmap_data,
                'weekdayPatterns': weekday_patterns,
                'peakTimes': peak_times_formatted,
                'legend': {
                    0: '매우 한산',
                    1: '한산',
                    2: '보통',
                    3: '혼잡',
                    4: '매우 혼잡'
                },
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            },
            message="혼잡도 히트맵 데이터를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"Congestion heatmap error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="혼잡도 히트맵 조회 중 오류가 발생했습니다.",
            code="HEATMAP_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def nfc_usage_analytics(request):
    """
    NFC 태그 사용률 통계 API - GET /analytics/nfc-usage
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 쿼리 파라미터
        start_date = request.GET.get('startDate', (timezone.now() - timedelta(days=30)).isoformat())
        end_date = request.GET.get('endDate', timezone.now().isoformat())
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # 전체 태그 통계
        total_tags = NFCTag.objects.count()
        active_tags = NFCTag.objects.filter(is_active=True).count()
        
        # 사용률 통계
        tag_usage = []
        for tag in NFCTag.objects.filter(is_active=True):
            logs = TagLog.objects.filter(
                tag=tag,
                timestamp__range=[start_date, end_date]
            )
            
            usage_data = {
                'tagId': str(tag.tag_id),
                'code': tag.code,
                'location': tag.get_location_display(),
                'totalScans': logs.count(),
                'uniqueUsers': logs.values('user').distinct().count(),
                'errorCount': logs.filter(action_type='error').count(),
                'lastScanTime': tag.last_scanned_at.isoformat() if tag.last_scanned_at else None
            }
            
            # 사용률 계산 (일평균 스캔 수)
            days_diff = (end_date - start_date).days or 1
            usage_data['dailyAverage'] = round(usage_data['totalScans'] / days_diff, 2)
            
            # 오류율 계산
            usage_data['errorRate'] = round(
                (usage_data['errorCount'] / usage_data['totalScans'] * 100)
                if usage_data['totalScans'] > 0 else 0, 2
            )
            
            tag_usage.append(usage_data)
        
        # 사용률 순위
        tag_usage_sorted = sorted(tag_usage, key=lambda x: x['totalScans'], reverse=True)
        
        # 위치별 집계
        location_stats = TagLog.objects.filter(
            timestamp__range=[start_date, end_date]
        ).values('tag__building', 'tag__floor', 'tag__room').annotate(
            total_scans=Count('log_id'),
            unique_users=Count('user', distinct=True),
            avg_daily_scans=Count('log_id') / days_diff
        )
        
        # 위치 정보를 문자열로 조합
        location_stats_formatted = []
        for stat in location_stats:
            location_stats_formatted.append({
                'location': f"{stat['tag__building']} {stat['tag__floor']}층 {stat['tag__room']}",
                'total_scans': stat['total_scans'],
                'unique_users': stat['unique_users'],
                'avg_daily_scans': stat['avg_daily_scans']
            })
        
        # 시간대별 사용 패턴
        hourly_usage = TagLog.objects.filter(
            timestamp__range=[start_date, end_date]
        ).annotate(
            hour=ExtractHour('timestamp')
        ).values('hour').annotate(
            scan_count=Count('log_id')
        ).order_by('hour')
        
        # 미사용 태그 감지
        unused_tags = []
        for tag in NFCTag.objects.filter(is_active=True):
            if not tag.last_scanned_at or tag.last_scanned_at < timezone.now() - timedelta(days=7):
                unused_tags.append({
                    'tagId': str(tag.tag_id),
                    'code': tag.code,
                    'location': tag.get_location_display(),
                    'lastScanTime': tag.last_scanned_at.isoformat() if tag.last_scanned_at else None,
                    'daysSinceLastScan': (timezone.now() - tag.last_scanned_at).days if tag.last_scanned_at else None
                })
        
        return APIResponse.success(
            data={
                'summary': {
                    'totalTags': total_tags,
                    'activeTags': active_tags,
                    'inactiveTags': total_tags - active_tags,
                    'totalScans': sum(t['totalScans'] for t in tag_usage),
                    'averageScansPerTag': round(sum(t['totalScans'] for t in tag_usage) / active_tags, 2) if active_tags > 0 else 0
                },
                'tagUsage': {
                    'top10': tag_usage_sorted[:10],
                    'bottom10': tag_usage_sorted[-10:] if len(tag_usage_sorted) > 10 else [],
                    'all': tag_usage_sorted
                },
                'locationStats': location_stats_formatted,
                'hourlyPattern': list(hourly_usage),
                'unusedTags': unused_tags,
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            },
            message="NFC 태그 사용률 통계를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"NFC usage analytics error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="NFC 태그 사용률 통계 조회 중 오류가 발생했습니다.",
            code="NFC_USAGE_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def identify_bottlenecks(request):
    """
    병목 구간 식별 API - GET /analytics/bottlenecks
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 병목 구간 식별 기준
        # 1. 대기시간이 긴 구간
        # 2. 같은 위치에서 반복 스캔이 많은 구간
        # 3. 대기열이 자주 정체되는 검사실
        
        # 대기시간 기준 병목 구간
        long_wait_exams = Exam.objects.annotate(
            # 'queue' 대신 'queues' 사용
            avg_wait=Avg('queues__estimated_wait_time'),
            total_patients=Count('queues'),
            # 'queue' 대신 'queues' 사용
            completion_rate=Count('queues', filter=Q(queues__state='completed')) * 100.0 / Count('queues')
        ).filter(
            avg_wait__gt=30  # 평균 대기시간 30분 초과
        ).order_by('-avg_wait')
        
        wait_bottlenecks = []
        for exam in long_wait_exams:
            wait_bottlenecks.append({
                'examId': exam.exam_id,
                # 'exam_name' 대신 'title' 사용 (Exam 모델에 exam_name 필드가 없으므로)
                'examName': exam.title, 
                'department': exam.department,
                'avgWaitTime': round(exam.avg_wait or 0, 2),
                'totalPatients': exam.total_patients,
                'completionRate': round(exam.completion_rate or 0, 2),
                'severity': 'high' if exam.avg_wait > 60 else 'medium'
            })
        
        # 반복 스캔 기준 병목 구간
        # 먼저 사용자별 스캔 횟수를 계산
        user_scans = TagLog.objects.values(
            'tag__building', 'tag__floor', 'tag__room', 'user'
        ).annotate(
            scan_count=Count('log_id')
        ).filter(
            scan_count__gt=3  # 같은 사용자가 같은 위치에서 3회 이상 스캔
        )
        
        # 위치별로 집계
        repeat_scan_locations = []
        location_data = {}
        
        for scan in user_scans:
            # 위치를 문자열로 조합
            location = f"{scan['tag__building']} {scan['tag__floor']}층 {scan['tag__room']}"
            if location not in location_data:
                location_data[location] = {
                    'users': set(),
                    'scan_counts': []
                }
            location_data[location]['users'].add(scan['user'])
            location_data[location]['scan_counts'].append(scan['scan_count'])
        
        for location, data in location_data.items():
            if data['scan_counts']:
                repeat_scan_locations.append({
                    'location': location,
                    'total_repeat_users': len(data['users']),
                    'avg_scans_per_user': sum(data['scan_counts']) / len(data['scan_counts'])
                })
        
        # 정렬
        repeat_scan_locations = sorted(
            repeat_scan_locations, 
            key=lambda x: x['total_repeat_users'], 
            reverse=True
        )
        
        scan_bottlenecks = []
        for loc in repeat_scan_locations:
            scan_bottlenecks.append({
                'location': loc['location'],
                'affectedUsers': loc['total_repeat_users'],
                'avgScansPerUser': round(loc['avg_scans_per_user'] or 0, 2),
                'severity': 'high' if loc['avg_scans_per_user'] > 5 else 'medium'
            })
        
        # 시간대별 병목 현상
        hourly_bottlenecks = []
        for hour in range(8, 18):  # 업무시간 8-18시
            hour_queues = Queue.objects.filter(
                # `joined_at` 필드 대신 `created_at` 필드를 사용
                created_at__hour=hour,
                created_at__gte=timezone.now() - timedelta(days=7)
            )
            
            avg_wait = hour_queues.aggregate(avg=Avg('estimated_wait_time'))['avg'] or 0
            if avg_wait > 45:  # 45분 이상이면 병목
                hourly_bottlenecks.append({
                    'hour': hour,
                    'avgWaitTime': round(avg_wait, 2),
                    'patientCount': hour_queues.count(),
                    'severity': 'high' if avg_wait > 60 else 'medium'
                })
        
        # 프로세스 병목 (상태 전환 지연)
        process_bottlenecks = []
        
        # called 상태에서 in_progress로 전환이 오래 걸리는 경우
        # MySQL 호환성을 위해 Python에서 시간 계산 수행
        now = timezone.now()
        slow_queues = Queue.objects.filter(
            state='called',
            called_at__lt=now - timedelta(minutes=15)
        ).select_related('exam')
        
        # Python에서 지연 시간 계산
        slow_transitions = []
        exam_delays = {}
        
        for queue in slow_queues:
            exam_title = queue.exam.title if queue.exam else 'Unknown'
            if queue.called_at:
                delay_minutes = (now - queue.called_at).total_seconds() / 60.0
                
                if exam_title not in exam_delays:
                    exam_delays[exam_title] = {'delays': [], 'count': 0}
                
                exam_delays[exam_title]['delays'].append(delay_minutes)
                exam_delays[exam_title]['count'] += 1
        
        # 평균 계산
        for exam_title, data in exam_delays.items():
            if data['delays']:
                slow_transitions.append({
                    'exam__title': exam_title,
                    'count': data['count'],
                    'avg_delay': sum(data['delays']) / len(data['delays'])
                })
        
        for trans in slow_transitions:
            if trans['count'] > 0:
                process_bottlenecks.append({
                    'type': 'call_to_progress_delay',
                    'examName': trans['exam__title'], # 'exam__exam_name' 대신 'exam__title' 사용
                    'affectedCount': trans['count'],
                    'avgDelay': round(trans['avg_delay'] or 0, 2),
                    'severity': 'high'
                })
        
        # 개선 제안
        recommendations = []
        
        if wait_bottlenecks:
            recommendations.append({
                'area': '대기시간',
                'issue': f"{len(wait_bottlenecks)}개 검사실에서 평균 대기시간이 30분을 초과합니다.",
                'suggestion': '해당 검사실의 인력 증원 또는 프로세스 개선이 필요합니다.'
            })
        
        if scan_bottlenecks:
            recommendations.append({
                'area': '동선',
                'issue': f"{len(scan_bottlenecks)}개 위치에서 환자들이 길을 찾지 못해 반복 스캔하고 있습니다.",
                'suggestion': '해당 위치의 안내 표지판 개선 또는 추가 안내 인력 배치가 필요합니다.'
            })
        
        if hourly_bottlenecks:
            peak_hours = [b['hour'] for b in hourly_bottlenecks]
            recommendations.append({
                'area': '시간대',
                'issue': f"{peak_hours}시에 병목 현상이 발생합니다.",
                'suggestion': '예약 시간 분산 또는 피크 시간대 추가 인력 배치가 필요합니다.'
            })
        
        return APIResponse.success(
            data={
                'waitTimeBottlenecks': wait_bottlenecks,
                'navigationBottlenecks': scan_bottlenecks,
                'hourlyBottlenecks': hourly_bottlenecks,
                'processBottlenecks': process_bottlenecks,
                'recommendations': recommendations,
                'summary': {
                    'totalBottlenecks': len(wait_bottlenecks) + len(scan_bottlenecks) + len(hourly_bottlenecks),
                    'highSeverity': len([b for b in wait_bottlenecks + scan_bottlenecks if b['severity'] == 'high']),
                    'mediumSeverity': len([b for b in wait_bottlenecks + scan_bottlenecks if b['severity'] == 'medium'])
                }
            },
            message="병목 구간을 식별했습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"Bottleneck identification error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="병목 구간 식별 중 오류가 발생했습니다.",
            code="BOTTLENECK_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def custom_report(request):
    """
    사용자 정의 보고서 생성 API - POST /analytics/custom-report
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 보고서 파라미터
        report_type = request.data.get('reportType', 'general')
        metrics = request.data.get('metrics', ['waitTime', 'patientFlow'])
        start_date = request.data.get('startDate', (timezone.now() - timedelta(days=30)).isoformat())
        end_date = request.data.get('endDate', timezone.now().isoformat())
        departments = request.data.get('departments', [])
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        report_data = {
            'metadata': {
                'reportType': report_type,
                'generatedAt': timezone.now().isoformat(),
                'generatedBy': request.user.username,
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            },
            'data': {}
        }
        
        # 요청된 메트릭별 데이터 수집
        if 'waitTime' in metrics:
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            wait_stats = Queue.objects.filter(
                created_at__range=[start_date, end_date]
            )
            if departments:
                wait_stats = wait_stats.filter(exam__department__in=departments)
            
            report_data['data']['waitTime'] = {
                'average': wait_stats.aggregate(avg=Avg('estimated_wait_time'))['avg'] or 0,
                'max': wait_stats.aggregate(max=Max('estimated_wait_time'))['max'] or 0,
                'trend': list(wait_stats.annotate(
                    # `joined_at` 필드 대신 `created_at` 필드를 사용
                    date=TruncDate('created_at')
                ).values('date').annotate(
                    avg_wait=Avg('estimated_wait_time')
                ).order_by('date'))
            }
        
        if 'patientFlow' in metrics:
            flow_data = TagLog.objects.filter(
                timestamp__range=[start_date, end_date]
            )
            
            popular_locations = flow_data.values(
                'tag__building', 'tag__floor', 'tag__room'
            ).annotate(
                count=Count('log_id')
            ).order_by('-count')[:10]
            
            # 위치 정보를 문자열로 조합
            popular_locations_formatted = []
            for loc in popular_locations:
                popular_locations_formatted.append({
                    'location': f"{loc['tag__building']} {loc['tag__floor']}층 {loc['tag__room']}",
                    'count': loc['count']
                })
            
            report_data['data']['patientFlow'] = {
                'totalScans': flow_data.count(),
                'uniquePatients': flow_data.values('user').distinct().count(),
                'popularLocations': popular_locations_formatted
            }
        
        if 'efficiency' in metrics:
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            completed = Queue.objects.filter(
                created_at__range=[start_date, end_date],
                state='completed'
            ).count()
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            total = Queue.objects.filter(
                created_at__range=[start_date, end_date]
            ).count()
            
            report_data['data']['efficiency'] = {
                'completionRate': round(completed / total * 100, 2) if total > 0 else 0,
                'totalProcessed': completed,
                'totalPatients': total
            }
        
        # 보고서 로그 저장
        AdminLog.objects.create(
            user=request.user,
            action='generate_report',
            details=json.dumps({
                'report_type': report_type,
                'metrics': metrics,
                'departments': departments
            })
        )
        
        return APIResponse.success(
            data=report_data,
            message="사용자 정의 보고서가 생성되었습니다.",
            status_code=status.HTTP_200_OK
        )
        
    except Exception as e:
        logger.error(f"Custom report generation error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="보고서 생성 중 오류가 발생했습니다.",
            code="REPORT_GENERATION_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def export_data(request):
    """
    데이터 내보내기 API - GET /analytics/export
    
    CSV/Excel/PDF 형식으로 데이터 내보내기
    """
    is_admin, error_msg = _check_analytics_permission(request)
    if not is_admin:
        return APIResponse.error(
            message=error_msg,
            code="FORBIDDEN",
            status_code=status.HTTP_403_FORBIDDEN
        )
    
    try:
        # 파라미터
        export_type = request.GET.get('type', 'csv')  # csv, excel, pdf
        data_type = request.GET.get('dataType', 'queue')  # queue, nfc, analytics
        start_date = request.GET.get('startDate', (timezone.now() - timedelta(days=30)).isoformat())
        end_date = request.GET.get('endDate', timezone.now().isoformat())
        
        # 날짜 변환
        start_date = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
        
        # 데이터 수집
        data, headers = _get_export_data(data_type, start_date, end_date)
        
        # 로그 저장
        AdminLog.objects.create(
            user=request.user,
            action='export_data',
            details=json.dumps({
                'export_type': export_type,
                'data_type': data_type,
                'period': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                }
            })
        )
        
        # 형식별 내보내기
        if export_type == 'csv':
            return _export_csv(data, headers, data_type)
        elif export_type == 'excel' and EXCEL_AVAILABLE:
            return _export_excel(data, headers, data_type)
        elif export_type == 'pdf' and PDF_AVAILABLE:
            return _export_pdf(data, headers, data_type, start_date, end_date)
        else:
            return APIResponse.error(
                message=f"{export_type} 형식은 지원되지 않습니다.",
                code="UNSUPPORTED_FORMAT",
                status_code=status.HTTP_400_BAD_REQUEST
            )
        
    except Exception as e:
        logger.error(f"Data export error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="데이터 내보내기 중 오류가 발생했습니다.",
            code="EXPORT_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

def _get_export_data(data_type, start_date, end_date):
    """내보낼 데이터 수집"""
    if data_type == 'queue':
        headers = ['Queue ID', 'Queue Number', 'Patient', 'Exam', 'State', 'Priority', 'Wait Time', 'Joined At']
        
        # `joined_at` 필드 대신 `created_at` 필드를 사용
        queues = Queue.objects.filter(
            created_at__range=[start_date, end_date]
        ).select_related('user', 'exam')
        
        data = []
        for q in queues:
            data.append([
                str(q.queue_id),
                q.queue_number,
                q.user.username if q.user else 'N/A',
                # `exam_name` 대신 `title` 사용
                q.exam.title if q.exam else 'N/A', 
                q.state,
                q.priority,
                q.estimated_wait_time or 0,
                # `joined_at` 필드 대신 `created_at` 필드를 사용
                q.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
    
    elif data_type == 'nfc':
        headers = ['Tag ID', 'Code', 'Location', 'Active', 'Total Scans', 'Last Scan']
        
        tags = NFCTag.objects.all()
        data = []
        
        for tag in tags:
            scan_count = TagLog.objects.filter(
                tag=tag,
                timestamp__range=[start_date, end_date]
            ).count()
            
            data.append([
                str(tag.tag_id),
                tag.code,
                tag.get_location_display(),
                'Yes' if tag.is_active else 'No',
                scan_count,
                tag.last_scanned_at.strftime('%Y-%m-%d %H:%M:%S') if tag.last_scanned_at else 'Never'
            ])
    
    elif data_type == 'analytics':
        headers = ['Date', 'Total Patients', 'Avg Wait Time', 'Completion Rate', 'Total Scans']
        data = []
        
        for date in range((end_date - start_date).days + 1):
            target_date = start_date + timedelta(days=date)
            
            # `joined_at` 필드 대신 `created_at` 필드를 사용
            day_stats = Queue.objects.filter(
                created_at__date=target_date.date()
            ).aggregate(
                total=Count('queue_id'),
                avg_wait=Avg('estimated_wait_time'),
                completed=Count('queue_id', filter=Q(state='completed'))
            )
            
            scans = TagLog.objects.filter(
                timestamp__date=target_date.date()
            ).count()
            
            completion_rate = (day_stats['completed'] / day_stats['total'] * 100) if day_stats['total'] > 0 else 0
            
            data.append([
                target_date.strftime('%Y-%m-%d'),
                day_stats['total'] or 0,
                round(day_stats['avg_wait'] or 0, 2),
                round(completion_rate, 2),
                scans
            ])
    
    return data, headers

def _export_csv(data, headers, data_type):
    """CSV 형식으로 내보내기"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{data_type}_export_{timezone.now().strftime("%Y%m%d")}.csv"'
    
    writer = csv.writer(response)
    writer.writerow(headers)
    
    for row in data:
        writer.writerow(row)
    
    return response

def _export_excel(data, headers, data_type):
    """Excel 형식으로 내보내기"""
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{data_type.upper()} Export"
    
    # 헤더 스타일 설정
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 헤더 작성
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 작성
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # 열 너비 자동 조정
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 메모리에서 파일 생성
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{data_type}_export_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    return response

def _export_pdf(data, headers, data_type, start_date, end_date):
    """PDF 형식으로 내보내기"""
    output = io.BytesIO()
    
    # PDF 문서 생성
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    # 제목 스타일
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # 중앙 정렬
    )
    
    # 제목 추가
    title = Paragraph(f"{data_type.upper()} Export Report", title_style)
    elements.append(title)
    
    # 기간 정보
    period_text = f"Period: {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
    period = Paragraph(period_text, styles['Normal'])
    elements.append(period)
    elements.append(Spacer(1, 20))
    
    # 요약 통계 (analytics인 경우)
    if data_type == 'analytics' and data:
        summary_data = [
            ['Metric', 'Value'],
            ['Total Records', str(len(data))],
            ['Date Range', f"{(end_date - start_date).days + 1} days"],
            ['Generated At', timezone.now().strftime('%Y-%m-%d %H:%M:%S')]
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 30))
    
    # 메인 데이터 테이블
    table_data = [headers] + data[:50]  # PDF는 최대 50행으로 제한
    
    # 테이블 생성
    table = Table(table_data)
    
    # 테이블 스타일 적용
    table.setStyle(TableStyle([
        # 헤더 스타일
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # 데이터 행 스타일
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        
        # 교대로 배경색 적용
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white])
    ]))
    
    elements.append(table)
    
    # 50행 초과 시 안내문
    if len(data) > 50:
        note = Paragraph(
            f"Note: Only first 50 records shown. Total records: {len(data)}",
            styles['Normal']
        )
        elements.append(Spacer(1, 20))
        elements.append(note)
    
    # PDF 생성
    doc.build(elements)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/pdf'
    )
    response['Content-Disposition'] = f'attachment; filename="{data_type}_export_{timezone.now().strftime("%Y%m%d")}.pdf"'

    return response


@api_view(['GET'])
@permission_classes([permissions.AllowAny])  # 테스트를 위해 임시로 AllowAny
def predictions(request):
    """
    LSTM 기반 대기 시간 및 혼잡도 예측 API
    GET /api/v1/analytics/predictions/
    """
    # 권한 확인 (테스트를 위해 임시 비활성화)
    # if not request.user.role or request.user.role not in ['super', 'dept']:
    #     return APIResponse.error(
    #         message="부서 관리자 이상의 권한이 필요합니다.",
    #         code="FORBIDDEN",
    #         status_code=status.HTTP_403_FORBIDDEN
    #     )

    try:
        logger.info("=" * 80)
        logger.info("=== predictions API 호출됨 ===")
        logger.info(f"Request method: {request.method}")
        logger.info(f"Request path: {request.path}")
        logger.info(f"Query params: {request.GET}")
        logger.info("=" * 80)

        # 1. 데모 모드 확인
        demo_active = cache.get("demo_mode_active")
        logger.info(f"데모 모드 활성: {demo_active}")

        if demo_active:
            try:
                # 데모 타임라인에서 현재 시점 데이터 가져오기
                start_time_str = cache.get("demo_start_time")
                timeline_json = cache.get("demo_timeline")

                if start_time_str and timeline_json:
                    timeline = json.loads(timeline_json)
                    start_time = datetime.fromisoformat(start_time_str)
                    current_time = timezone.now()

                    # 타임존 정보 맞추기
                    if start_time.tzinfo is None:
                        start_time = timezone.make_aware(start_time)

                    elapsed = int((current_time - start_time).total_seconds())

                    # 10초 단위로 반올림
                    timeline_key = str((elapsed // 10) * 10)

                    logger.info(f"데모 경과 시간: {elapsed}초, timeline_key: {timeline_key}")

                    # 타임라인에서 해당 시점 데이터 찾기
                    if timeline_key in timeline:
                        demo_predictions = timeline[timeline_key]
                        logger.info(f"데모 데이터 발견: {len(demo_predictions)}개 부서")

                        # 전체 혼잡도 계산
                        all_congestions = [
                            dept_data.get('congestion', 0)
                            for dept_data in demo_predictions.values()
                            if isinstance(dept_data, dict) and 'congestion' in dept_data
                        ]
                        overall_congestion = sum(all_congestions) / len(all_congestions) if all_congestions else 0

                        # 데모 데이터 반환
                        return APIResponse.success(
                            data={
                                'timestamp': timezone.now().isoformat(),
                                'timeframe': request.GET.get('timeframe', '30min'),
                                'demo_mode': True,
                                'demo_elapsed': elapsed,
                                'overall': {
                                    'avgCongestion': round(overall_congestion, 2),
                                    'congestionLevel': _get_congestion_level(overall_congestion),
                                    'totalDepartments': len(demo_predictions)
                                },
                                'departments': demo_predictions,
                                'recommendations': [
                                    {
                                        'type': 'demo',
                                        'message': f'데모 모드 실행 중 ({elapsed}초 경과)',
                                        'action': '실시간 변화를 관찰하세요'
                                    }
                                ] + _generate_recommendations(demo_predictions)
                            },
                            message="데모 데이터",
                            status_code=status.HTTP_200_OK
                        )
                else:
                    logger.warning("데모 모드 활성이지만 타임라인 데이터 없음")
            except Exception as demo_error:
                logger.error(f"데모 모드 처리 중 오류: {demo_error}", exc_info=True)
                # 데모 실패 시 일반 모드로 폴백

        # 2. 일반 모드 - 예측 서비스 호출
        timeframe = request.GET.get('timeframe', '30min')
        department = request.GET.get('department')

        logger.info(f"일반 모드 예측 시작: timeframe={timeframe}, department={department}")

        # 예측 서비스 호출 (에러 핸들링 추가)
        try:
            prediction_data = PredictionService.get_predictions(timeframe=timeframe)
            logger.info(f"예측 서비스 성공: {len(prediction_data)}개 부서")
        except Exception as pred_error:
            logger.error(f"PredictionService 오류: {pred_error}", exc_info=True)
            # 빈 데이터로 폴백 (500 에러 대신)
            prediction_data = {}
            logger.warning("빈 예측 데이터로 폴백")

        # 특정 부서만 필터링
        if department and department in prediction_data:
            prediction_data = {department: prediction_data[department]}

        # 전체 혼잡도 계산
        all_congestions = [
            dept_data.get('congestion', 0)
            for dept_data in prediction_data.values()
            if isinstance(dept_data, dict) and 'congestion' in dept_data
        ]
        overall_congestion = sum(all_congestions) / len(all_congestions) if all_congestions else 0

        # 응답 데이터 구성
        response_data = {
            'timestamp': timezone.now().isoformat(),
            'timeframe': timeframe,
            'overall': {
                'avgCongestion': round(overall_congestion, 2),
                'congestionLevel': _get_congestion_level(overall_congestion),
                'totalDepartments': len(prediction_data)
            },
            'departments': prediction_data,
            'recommendations': _generate_recommendations(prediction_data)
        }

        logger.info(f"응답 데이터 생성 완료: {len(prediction_data)}개 부서")
        logger.info("=" * 80)

        return APIResponse.success(
            data=response_data,
            message="대기시간 예측 데이터를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )

    except Exception as e:
        logger.error(f"Prediction API 최상위 에러: {str(e)}", exc_info=True)
        logger.error(f"에러 타입: {type(e).__name__}")
        logger.error(f"에러 상세: {e}")

        # 500 에러 대신 200 OK + 빈 데이터 반환 (Frontend가 처리 가능하도록)
        return APIResponse.success(
            data={
                'timestamp': timezone.now().isoformat(),
                'timeframe': request.GET.get('timeframe', '30min'),
                'overall': {
                    'avgCongestion': 0,
                    'congestionLevel': 'unknown',
                    'totalDepartments': 0
                },
                'departments': {},
                'recommendations': [],
                'error': str(e)  # 디버깅용
            },
            message=f"예측 데이터 조회 중 오류가 발생했습니다: {str(e)}",
            status_code=status.HTTP_200_OK
        )


def _get_congestion_level(congestion_value):
    """혼잡도 수치를 레벨로 변환"""
    if congestion_value < 0.2:
        return 'very_low'
    elif congestion_value < 0.4:
        return 'low'
    elif congestion_value < 0.6:
        return 'moderate'
    elif congestion_value < 0.8:
        return 'high'
    else:
        return 'very_high'


def _generate_recommendations(prediction_data):
    """예측 데이터 기반 추천 생성"""
    recommendations = []

    # 혼잡도가 높은 부서 찾기
    congested_depts = [
        dept for dept, data in prediction_data.items()
        if isinstance(data, dict) and data.get('congestion', 0) > 0.7
    ]

    if congested_depts:
        recommendations.append({
            'type': 'congestion',
            'message': f"{', '.join(congested_depts)} 부서가 혼잡할 것으로 예상됩니다.",
            'action': '추가 인력 배치 또는 환자 분산을 고려하세요.'
        })

    return recommendations


@api_view(['GET'])
@permission_classes([])  # 테스트를 위해 임시로 권한 제거
def predictions_timeline(request):
    """
    시계열 예측 데이터 API
    GET /api/v1/analytics/predictions/timeline/
    """
    try:
        timeline_data = PredictionService.get_timeline_predictions()

        return APIResponse.success(
            data={
                'timestamp': timezone.now().isoformat(),
                'timeline': timeline_data
            },
            message="시계열 예측 데이터를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )

    except Exception as e:
        logger.error(f"Timeline prediction error: {str(e)}", exc_info=True)
        return APIResponse.error(
            message="시계열 예측 데이터 조회 중 오류가 발생했습니다.",
            code="TIMELINE_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([])  # 테스트를 위해 임시로 권한 제거
def predictions_domino(request):
    """
    도미노 효과 예측 API
    POST /api/v1/analytics/predictions/domino/
    """
    logger.info("=" * 80)
    logger.info("=== predictions_domino 호출됨 ===")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request path: {request.path}")
    logger.info(f"Request data: {request.data}")
    logger.info(f"Content-Type: {request.content_type}")
    logger.info("=" * 80)

    try:
        source_dept = request.data.get('source_department')
        delay_minutes = request.data.get('delay_minutes', 30)

        logger.info(f"Parsed params: source_dept={source_dept}, delay_minutes={delay_minutes}")

        if not source_dept:
            return APIResponse.error(
                message="소스 부서를 지정해주세요.",
                code="MISSING_PARAMETER",
                status_code=status.HTTP_400_BAD_REQUEST
            )

        domino_data = PredictionService.get_domino_predictions(source_dept, delay_minutes)

        logger.info(f"Domino prediction 결과 개수: {len(domino_data)}")
        logger.info(f"Domino data 샘플: {domino_data[:2] if domino_data else 'No data'}")

        # Frontend가 배열을 직접 기대하므로 impacts만 반환
        result = APIResponse.success(
            data=domino_data,  # 배열 직접 반환
            message="도미노 효과 예측을 완료했습니다.",
            status_code=status.HTTP_200_OK
        )

        logger.info(f"Response data 키: {result.data.keys() if hasattr(result, 'data') else 'No data attr'}")
        return result

    except Exception as e:
        logger.error(f"Domino prediction error: {str(e)}", exc_info=True)
        logger.error(f"Request data: source_dept={request.data.get('source_department')}, delay={request.data.get('delay_minutes')}")
        return APIResponse.error(
            message=f"도미노 효과 예측 중 오류가 발생했습니다: {str(e)}",
            code="DOMINO_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_classes([])  # 테스트를 위해 임시로 권한 제거
def predictions_heatmap(request):
    """
    혼잡도 히트맵 예측 API
    GET /api/v1/analytics/predictions/heatmap/
    """
    logger.info("=" * 80)
    logger.info("=== predictions_heatmap 호출됨 ===")
    logger.info(f"Request method: {request.method}")
    logger.info(f"Request path: {request.path}")
    logger.info("=" * 80)

    try:
        heatmap_data = PredictionService.get_heatmap_predictions()

        logger.info(f"Heatmap 데이터 개수: {len(heatmap_data)}")
        logger.info(f"Heatmap data 샘플: {heatmap_data[:3] if heatmap_data else 'No data'}")

        result = APIResponse.success(
            data=heatmap_data,  # Frontend가 기대하는 구조로 직접 반환
            message="혼잡도 히트맵 데이터를 조회했습니다.",
            status_code=status.HTTP_200_OK
        )

        logger.info(f"Response 생성 완료")
        return result

    except Exception as e:
        logger.error(f"Heatmap prediction error: {str(e)}", exc_info=True)
        logger.error(f"Heatmap prediction error details", exc_info=True)
        return APIResponse.error(
            message=f"히트맵 데이터 조회 중 오류가 발생했습니다: {str(e)}",
            code="HEATMAP_ERROR",
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

    # 대기시간이 긴 부서
    long_wait_depts = [
        dept for dept, data in prediction_data.items()
        if isinstance(data, dict) and data.get('predicted_wait', 0) > 60
    ]

    if long_wait_depts:
        recommendations.append({
            'type': 'wait_time',
            'message': f"{', '.join(long_wait_depts)} 부서의 예상 대기시간이 60분을 초과합니다.",
            'action': '프로세스 개선 또는 예약 시간 조정을 검토하세요.'
        })

    # 상승 추세 부서
    rising_depts = [
        dept for dept, data in prediction_data.items()
        if isinstance(data, dict) and data.get('trend') == 'up'
    ]

    if rising_depts:
        recommendations.append({
            'type': 'trend',
            'message': f"{', '.join(rising_depts)} 부서의 대기시간이 증가 추세입니다.",
            'action': '사전 대응 조치를 준비하세요.'
        })

    return recommendations